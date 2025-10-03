# Den_review.py — 5KB byte-based chunking + 7-column DF + readable Excel
# Python 3.12 compatible

from __future__ import annotations

import io
import re
from copy import copy
from typing import Any, Dict, List, Tuple

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from dotenv import load_dotenv

# These are your helpers/modules
import format_document_data
import format_dataframe
import re
from mdr.client.llm import LLMClient


# ----------------------- Chunking (bytes, ~5KB) -----------------------

HEADER_RE = re.compile(r'^(#{1,6})\s*(.+?)\s*$', re.MULTILINE)

def parse_headers(markdown_txt: str):
    """
    Returns list of (start_index, title) for all Markdown headers.
    """
    headers = []
    for m in HEADER_RE.finditer(markdown_txt):
        headers.append((m.start(), m.group(2)))
    headers.sort(key=lambda x: x[0])
    return headers

def nearest_header_for(markdown_txt: str, headers, snippet: str) -> str:
    """
    Finds the nearest preceding header for a snippet of text.
    """
    if not snippet:
        return ""
    probe = snippet.strip().replace("\n", " ")[:60]  # short search key
    pos = markdown_txt.find(probe)
    if pos == -1:
        return ""
    chosen = ""
    for start, title in headers:
        if start <= pos:
            chosen = title
        else:
            break
    return chosen

def _bytes_len(s: str) -> int:
    return len(s.encode("utf-8", errors="ignore"))

def smart_chunk_bytes(text: str, max_kb: int = 5, soft_min_kb: int = 4) -> List[str]:
    """
    Chunk a long string into ~max_kb (UTF-8 bytes) pieces, preferring paragraph boundaries,
    with a sentence/punctuation-aware fallback if a single paragraph is huge.

    - max_kb: hard cap per chunk in kilobytes
    - soft_min_kb: if a unit is very long, try to cut after this many KB at a natural boundary
    """
    if not text:
        return [""]

    max_bytes = max_kb * 1024
    soft_min_bytes = soft_min_kb * 1024

    # If text already small, single chunk
    if _bytes_len(text) <= max_bytes:
        return [text]

    # Split paragraphs but keep separators so structure survives
    parts = re.split(r"(\n\s*\n)", text)  # keeps the double-newline separators
    chunks: List[str] = []
    buf: List[str] = []
    buf_bytes = 0

    def flush():
        nonlocal buf, buf_bytes, chunks
        if buf:
            chunk = "".join(buf).strip()
            if chunk:
                chunks.append(chunk)
            buf, buf_bytes = [], 0

    for part in parts:
        b_len = _bytes_len(part)
        # If it fits, accumulate
        if buf_bytes + b_len <= max_bytes:
            buf.append(part)
            buf_bytes += b_len
            continue

        # If the part itself is massive, split it with sentence-aware fallback
        if b_len > max_bytes:
            # flush current buffer first
            flush()
            start = 0
            # Work on the raw string by indexes, but measure slices in bytes
            while start < len(part):
                # optimistic end bound by characters; we adjust by bytes
                # Expand window to near max_bytes (by bytes)
                end_char = min(len(part), start + max_bytes)  # rough char bound
                # tighten end_char so bytes <= max_bytes
                # (binary-shrink to avoid quadratic worst-case)
                lo, hi = start, end_char
                best = start
                while lo <= hi:
                    mid = (lo + hi) // 2
                    if _bytes_len(part[start:mid]) <= max_bytes:
                        best = mid
                        lo = mid + 1
                    else:
                        hi = mid - 1
                end_char = best

                # If we have room to look for a natural boundary and the slice is reasonably large
                slice_str = part[start:end_char]
                if _bytes_len(slice_str) >= soft_min_bytes:
                    # try to land on punctuation or whitespace near the tail
                    # prioritize hard breaks/new paragraphs, then sentence-ish, then spaces
                    tail_window = slice_str[-min(len(slice_str), 400):]  # last ~400 chars
                    cut_rel = max(
                        tail_window.rfind("\n\n"),
                        tail_window.rfind("。\n"),
                        tail_window.rfind("。」"),
                        tail_window.rfind(". "),
                        tail_window.rfind("。"),
                        tail_window.rfind("、"),
                        tail_window.rfind("，"),
                        tail_window.rfind(", "),
                        tail_window.rfind(" "),
                    )
                    if cut_rel > 10:  # avoid cutting too close to start
                        # move end_char back to that boundary
                        end_char = end_char - (len(tail_window) - cut_rel - 1)
                        slice_str = part[start:end_char]

                chunks.append(slice_str.strip())
                start = end_char
        else:
            # Current part would overflow the buffer; flush and then add to new buffer
            flush()
            buf.append(part)
            buf_bytes = b_len

    flush()
    # remove empties
    return [c for c in chunks if c]


# ----------------------- Utilities -----------------------

def sanitize_string(value: Any) -> Any:
    if isinstance(value, str):
        # Remove ASCII control chars that upset Excel cells
        return re.sub(r"[\x00-\x1F]", "", value)
    return value


def enforce_7_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Your logs showed 8 columns including 'result'. The sheet should have 7.
    Drop 'result' if present and log the final column list.
    """
    if "result" in df.columns:
        df = df.drop(columns=["result"])
    print(df.columns, "Number of columns in df which should be 7")
    return df


def apply_readability(ws: openpyxl.worksheet.worksheet.Worksheet, df: pd.DataFrame):
    """
    Make Excel legible:
      - Wrap text for long text columns like 'target_chunk' and 'AI_chunk'
      - Set column widths so the first column isn't a wall of text
    """
    col_widths = []
    for cname in df.columns:
        if cname in ("All_No", "pointout_level", "pointout_category"):
            col_widths.append(12)
        elif cname in ("point_why", "point_imp"):
            col_widths.append(28)
        elif cname in ("target_chunk", "AI_chunk"):
            col_widths.append(60)
        else:
            col_widths.append(24)

    for idx, width in enumerate(col_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    # Wrap text + align top for all data cells we just wrote (start at row 2 per writer)
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

def build_export_df(markdown_txt: str, final_rows):
    """
    Export 7 columns. チェック箇所 is auto-filled from nearest header in markdown_txt.
    """
    import pandas as pd
    headers = parse_headers(markdown_txt)

    rows = []
    for r in final_rows:
        if not isinstance(r, dict):
            continue

        snippet = r.get("target_chunk") or r.get("AI_chunk") or ""
        check_spot = (
            r.get("section")
            or r.get("chapter")
            or r.get("chapter_title")
            or nearest_header_for(markdown_txt, headers, snippet)
        )

        rows.append({
            "No": r.get("All_No", ""),
            "チェック箇所": check_spot,
            "指摘箇所": snippet,
            "指摘理由": r.get("point_why", ""),
            "改善提案": r.get("point_imp", ""),
            "処置難易度": r.get("pointout_level", ""),
            "指摘分類": r.get("pointout_category", ""),
        })

    return pd.DataFrame(rows, columns=[
        "No", "チェック箇所", "指摘箇所", "指摘理由", "改善提案", "処置難易度", "指摘分類"
    ])

def write_export_df_to_sheet(ws, export_df):
    # Writes starting at row 2, using fixed columns:
    #  A No | B チェック箇所 | C 指摘箇所 | D 指摘理由 | E 改善提案 | I 処置難易度 | J 指摘分類
    from openpyxl.styles import Alignment
    col_map = {
        "No": 1,          # A
        "チェック箇所": 2,  # B
        "指摘箇所": 3,    # C
        "指摘理由": 4,    # D
        "改善提案": 5,    # E
        "処置難易度": 9,  # I
        "指摘分類": 10,   # J
    }
    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
        for col_name, excel_col in col_map.items():
            cell = ws.cell(row=row_idx, column=excel_col, value=row[col_name])
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def save_to_excel(
    export_df,                     # NOTE: this is the 7-column df from build_export_df
    template_path: str,
    specific_strings_json: dict,
    before_sheet_name: str = "レビュー結果_除外前",
    main_sheet_name: str = "レビュー結果",
):
    import io, openpyxl
    from copy import copy
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    output_excel = io.BytesIO()
    wb = openpyxl.load_workbook(template_path)

    # If you want to keep your existing filter step, apply it to the export df:
    filtered_df = format_dataframe.filter_dataframe(export_df, specific_strings_json)

    # Optional: dump the raw export df to the "除外前" sheet (columns laid out sequentially)
    if before_sheet_name in wb.sheetnames:
        ws_before = wb[before_sheet_name]
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                base_cell = ws_before.cell(row=2, column=c_idx)
                new_cell = ws_before.cell(row=r_idx, column=c_idx, value=sanitize_string(value))
                if base_cell.has_style:
                    new_cell._style = copy(base_cell._style)

    # Main target sheet (we write into A,B,C,D,E,I,J)
    if main_sheet_name not in wb.sheetnames:
        wb.create_sheet(main_sheet_name)
    ws = wb[main_sheet_name]

    # Clear previous data rows if any (keep header row 1 and styling in row 2 if you use it)
    max_existing = ws.max_row
    if max_existing > 2:
        ws.delete_rows(3, max_existing - 2)

    # Precise placement
    write_export_df_to_sheet(ws, filtered_df)

    # Optional: set widths to make it readable
    widths = {1: 8, 2: 24, 3: 60, 4: 36, 5: 36, 9: 14, 10: 14}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    pointout_count = len(filtered_df)
    wb.save(output_excel)
    output_excel.seek(0)
    return output_excel, pointout_count


# ----------------------- Main entry -----------------------

def create_AI_review(
    system_message_chat_conversation: str,
    markdown_txt: str,
    *,
    chunk_size_kb: int = 5,              # << byte-based target chunk size
    excel_template: str = "Den_temp.xlsx",
    specific_strings_json: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    """
    Two-phase review:
      1) Chunk the WHOLE markdown into ~chunk_size_kb (bytes) and review per chunk
      2) Section-based review (format_document_data.Section_chunks)
    Aggregates rows, numbers them, builds DataFrame, exports Excel.

    Returns:
      {
        "output_excel": BytesIO,
        "tokens": {"total_tokens": int, "prompt_tokens": int, "completion_tokens": int},
        "pointout_count": int
      }
    """
    print("create_AI_review function was called in Den_review file")
    load_dotenv()

    if specific_strings_json is None:
        specific_strings_json = {"domain_terms": ["半角カナ", "全角カナ"]}

    # -------- Phase 0: robust byte-based chunking (~5KB) --------
    try:
        text_chunks = smart_chunk_bytes(markdown_txt, max_kb=chunk_size_kb, soft_min_kb=max(1, chunk_size_kb - 1))
    except Exception as e:
        print("Error in smart_chunk_bytes; falling back to a single chunk:", e)
        text_chunks = [markdown_txt]

    print("DEBUG: len(text_chunks) =", len(text_chunks))

    # -------- Phase 1: per-chunk LLM review --------
    all_pre_rows: List[Dict[str, Any]] = []
    AllNo = 1

    pre_total_tokens = 0
    pre_prompt_tokens = 0
    pre_completion_tokens = 0

    x = 1
    for text_chunk in text_chunks:
        try:
            llm = LLMClient(model="gpt-4.1", temperature=0.0)
            data = llm.generate_review(system_message_chat_conversation, text_chunk)

            # IMPORTANT: correctly unpack tuple (rows_list, meta_dict)
            rows, _meta = llm.format_responses_with_meta(
                result_payload=data,
                all_no=AllNo,
                chapter_titles=text_chunk,
            )

            if isinstance(rows, list):
                all_pre_rows.extend(rows)
                AllNo += len(rows)
            else:
                print("format_responses_with_meta returned non-list rows; skipping.")

            toks = data.get("tokens", {}) if isinstance(data, dict) else {}
            pre_total_tokens += int(toks.get("total_tokens", 0) or 0)
            pre_prompt_tokens += int(toks.get("prompt_tokens", 0) or 0)
            pre_completion_tokens += int(toks.get("completion_tokens", 0) or 0)

        except Exception as e:
            print("Exception in the create_AI_review loop (Den_review.py):", e)

        print("x is -------------------------------", x)
        x += 1

    # -------- Phase 2: section-based review --------
    try:
        section_result = format_document_data.Section_chunks(
            markdown_txt, system_message_chat_conversation
        )
        if not isinstance(section_result, dict):
            raise ValueError("Section_chunks returned non-dict.")
    except Exception as e:
        print("error in the Section_chunks function:", e)
        section_result = {"all_responses": [], "tokens": {}}

    section_responses = section_result.get("all_responses", [])
    if not isinstance(section_responses, list):
        print("Warning: section_responses is not a list; forcing empty.")
        section_responses = []

    tokens_info = section_result.get("tokens", {})
    if not isinstance(tokens_info, dict):
        tokens_info = {}

    total_tokens = pre_total_tokens + int(tokens_info.get("section_total_tokens", 0) or 0)
    prompt_tokens = pre_prompt_tokens + int(tokens_info.get("section_prompt_tokens", 0) or 0)
    completion_tokens = pre_completion_tokens + int(tokens_info.get("section_completion_tokens", 0) or 0)

    # Continue numbering for section rows
    last_no = all_pre_rows[-1]["All_No"] if all_pre_rows and "All_No" in all_pre_rows[-1] else 0
    for r in section_responses:
        if isinstance(r, dict):
            last_no += 1
            r["All_No"] = last_no

    final_rows: List[Dict[str, Any]] = all_pre_rows + [r for r in section_responses if isinstance(r, dict)]

    print(type(final_rows), "Type of final_responses:\n")
    print("Final responses (first 3 rows):\n", final_rows[:3], "... total:", len(final_rows))

    # -------- DataFrame build (defensive) --------
    try:
        df = format_dataframe.custom_dataframe(final_rows)
        if not isinstance(df, pd.DataFrame):
            raise TypeError("custom_dataframe did not return a pandas DataFrame.")
    except Exception as e:
        print("Error in custom_dataframe function:", e)
        df = pd.DataFrame(final_rows)

    # Ensure exactly 7 columns for the template
    df = enforce_7_columns(df)

    # -------- Excel export (defensive) --------
    # After you have final_rows
    export_df = build_export_df(markdown_txt, final_rows)

    try:
        output_excel, pointout_count = save_to_excel(
            export_df,                     # << use export_df here
            template_path=excel_template,
            specific_strings_json=specific_strings_json,
        )
    except Exception as e:
        print("Error in the save_to_excel function:", e)
        import io
        output_excel, pointout_count = io.BytesIO(), 0

    print("create_AI_review function ended")

    return {
        "output_excel": output_excel,
        "tokens": {
            "total_tokens": total_tokens,
            "prompt_tokens": prompt_tokens,
            "completion_tokens": completion_tokens,
        },
        "pointout_count": pointout_count,
    }
